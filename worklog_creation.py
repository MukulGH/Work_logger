import os
import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import NamedStyle, PatternFill, Font, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

def get_previous_tuesday():
    today = datetime.date.today()
    if today.weekday() == 1:  # If today is Tuesday, return today
        return today
    else:
        # Calculate the difference between today and the previous Tuesday
        delta = (today.weekday() - 1) % 7
        # If today is Monday, delta will be 6, so we need to adjust it to get to the previous Tuesday
        delta = delta + 7 if delta == 6 else delta
        last_tuesday = today - datetime.timedelta(days=delta)
        return last_tuesday
def format_date(date):
    return date.strftime("%m_%d_%Y")

def get_filtered_categories(template_path):
    wb = load_workbook(template_path)
    ws = wb.active
    categories = [ws[f'A{i}'].value for i in range(6, 39) if ws[f'A{i}'].value]
    excluded_categories = [
        "GRT Product Development", "Product Upgrades", "Customer Projects",
        "Process Upgrades", "Production Build and Test", "Marketing"
    ]
    filtered_categories = [category for category in categories if category not in excluded_categories]
    return filtered_categories

def create_file(path, date_str, categories):
    filename = os.path.join(path, f"MG_worklog_week_of_{date_str}.xlsx")
    if os.path.exists(filename):
        raise FileExistsError("File already exists.")
    
    workbook = Workbook()
    headers = ["Start Time", "End Time", "Task", "Number of Hours", "Work Eligible", "Account"]
    
    # Defining styles
    header_fill = PatternFill(start_color="FFA07A", end_color="FFA07A", fill_type="solid")
    header_font = Font(bold=True, size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    dv = DataValidation(type="list", formula1="$K$4:$K$30", allow_blank=True)
    previous_tuesday = get_previous_tuesday()

    for i in range(7):
        day = format_date(previous_tuesday + datetime.timedelta(days=i))
        if i == 0:
            sheet = workbook.active
            sheet.title = day
        else:
            sheet = workbook.create_sheet(title=day)
        
        for col_num, header in enumerate(headers, start=2):  
            cell = sheet.cell(row=2, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        dv.add(f'G3:G22')
        sheet.add_data_validation(dv)
        
        for row_num in range(3, 23):
            if row_num == 3:
                continue
            sheet[f'B{row_num}'].number_format = 'hh:mm'
            sheet[f'C{row_num}'].number_format = 'hh:mm'
            sheet[f'E{row_num}'].number_format = 'hh:mm'
            # Setting default values
            sheet[f'B{row_num}'].value = datetime.time(5, 0)  # Default to 05:00 for Start Time
            sheet[f'C{row_num}'].value = datetime.time(0, 0)  # Default to 00:00 for End Time
            sheet[f'E{row_num}'].value = datetime.time(0, 0)  # Default to 00:00 for Number of Hours
            sheet[f'B{row_num}'] = f'=C{row_num-1}'
            sheet[f'C{row_num}'] = f'=B{row_num}+E{row_num}'
        
        sheet['I24'] = "Total hours worked today"
        sheet["I25"] = f'=SUM(E3:E22)'
        
        for i, category in enumerate(categories, start=4):
            sheet[f'K{i}'] = category
        
        # Adjusting column widths
        column_widths = [2, 12, 12, 30, 15, 15, 15]
        for i, width in enumerate(column_widths):
            sheet.column_dimensions[chr(65+i)].width = width
    
    workbook.save(filename)

if __name__ == "__main__":
    dir_path = os.getcwd()
    previous_tuesday = get_previous_tuesday()
    template_path = os.path.join(dir_path, "Timesheet_template_2023.xlsx")
    categories = get_filtered_categories(template_path)
    create_file(dir_path, format_date(previous_tuesday), categories)
