import os
import glob
import datetime
from openpyxl import load_workbook


def get_latest_excel_file(directory):
    excel_files = glob.glob(os.path.join(directory, '*.xlsx'))
    excel_files.sort(key=os.path.getmtime, reverse=True)
    return excel_files[0] if excel_files else None


def get_account_categories(file_path, sheet_name):
    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]

    categories = []
    for row in ws.iter_rows(min_row=3, min_col=11, max_row=30, max_col=11):
        for cell in row:
            if cell.value:
                categories.append(cell.value)

    return categories


def get_start_time_from_sheet(file_path, sheet_name):
    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]

    # Finding the next empty row in the "Task" column
    for row_num, row_cells in enumerate(ws.iter_rows(min_row=4, min_col=4, max_col=4), start=4):
        if row_cells[0].value is None:
            next_row_number = row_num
            break

    # Getting the start time from column B in the row signified by next_row_number
    start_time_cell = ws[f'B{next_row_number}']
    start_time = start_time_cell.value if start_time_cell.value else "00:00"

    return start_time


def update_worklog(file_path, sheet_name, start_time, task, hours, work_eligible, account):
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    # Finding the next empty row in the "Task" column
    for row_num, row_cells in enumerate(ws.iter_rows(min_row=4, min_col=4, max_col=4), start=4):
        if row_cells[0].value is None:
            next_row = row_num
            break

    # Updating the Excel sheet with the user inputs
    ws[f'B{next_row}'] = start_time
    ws[f'D{next_row}'] = task
    ws[f'E{next_row}'] = hours
    ws[f'F{next_row}'] = work_eligible
    ws[f'G{next_row}'] = account

    wb.save(file_path)


if __name__ == "__main__":
    directory = os.getcwd()
    latest_excel_file = get_latest_excel_file(directory)

    if latest_excel_file:
        today = datetime.date.today()
        sheet_name = today.strftime("%m_%d_%Y")
        wb = load_workbook(latest_excel_file, read_only=True)

        if sheet_name in wb.sheetnames:
            start_time = get_start_time_from_sheet(latest_excel_file, sheet_name)
            print(f"Start Time: {start_time}")
        else:
            print(f"Sheet '{sheet_name}' not found in the Excel file.")
    else:
        print("No Excel files found in the directory.")
